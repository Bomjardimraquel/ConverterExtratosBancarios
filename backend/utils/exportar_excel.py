import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List


def gerar_excel(lancamentos: List[dict], nome_empresa: str = "", banco: str = "", mes_ano: str = "") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Lançamentos"

    # ── Cores ────────────────────────────────────────────────────────────────
    COR_HEADER = "7B4FA6"
    COR_CREDITO = "F3EEFF"
    COR_DEBITO = "EDE7F6"
    COR_REVISAO = "FFEBEE"
    COR_TEXTO_HEADER = "FFFFFF"

    borda = Border(
        left=Side(style="thin", color="D1C4E9"),
        right=Side(style="thin", color="D1C4E9"),
        top=Side(style="thin", color="D1C4E9"),
        bottom=Side(style="thin", color="D1C4E9"),
    )

    # ── Título ───────────────────────────────────────────────────────────────
    titulo = f"Extrato Bancário - {banco}"
    if nome_empresa:
        titulo += f" | {nome_empresa}"
    if mes_ano:
        titulo += f" | {mes_ano}"

    ws.merge_cells("A1:F1")
    ws["A1"] = titulo
    ws["A1"].font = Font(bold=True, size=13, color=COR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Cabeçalho ────────────────────────────────────────────────────────────
    headers = ["Data", "Descrição", "Tipo", "Débito", "Crédito", "Valor (R$)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=COR_TEXTO_HEADER, size=10)
        cell.fill = PatternFill("solid", fgColor=COR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = borda
    ws.row_dimensions[2].height = 22

    # ── Dados ────────────────────────────────────────────────────────────────
    for row_idx, lanc in enumerate(lancamentos, 3):
        tipo = lanc.get("tipo", "")
        requer_revisao = lanc.get("requer_revisao", False)

        if requer_revisao:
            fill = PatternFill("solid", fgColor=COR_REVISAO)
        elif tipo == "Crédito":
            fill = PatternFill("solid", fgColor=COR_CREDITO)
        else:
            fill = PatternFill("solid", fgColor=COR_DEBITO)

        valores_linha = [
            lanc.get("data", ""),
            lanc.get("descricao", ""),
            tipo,
            lanc.get("conta_debito", ""),
            lanc.get("conta_credito", ""),
            lanc.get("valor", 0),
        ]

        for col, val in enumerate(valores_linha, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = borda
            cell.alignment = Alignment(vertical="center")
            if col == 6:  # Valor
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif col in (4, 5):  # Contas
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col == 1:  # Data
                cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row_idx].height = 18

    # ── Larguras das colunas ─────────────────────────────────────────────────
    larguras = [10, 55, 10, 12, 12, 15]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    # ── Linha de total ───────────────────────────────────────────────────────
    total_row = len(lancamentos) + 3
    ws.merge_cells(f"A{total_row}:E{total_row}")
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(bold=True, color=COR_TEXTO_HEADER)
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=COR_HEADER)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws[f"A{total_row}"].border = borda

    total_cell = ws.cell(row=total_row, column=6, value=f"=SUM(F3:F{total_row - 1})")
    total_cell.font = Font(bold=True, color=COR_TEXTO_HEADER)
    total_cell.fill = PatternFill("solid", fgColor=COR_HEADER)
    total_cell.number_format = '#,##0.00'
    total_cell.alignment = Alignment(horizontal="right", vertical="center")
    total_cell.border = borda
    ws.row_dimensions[total_row].height = 22

    # ── Freeze panes ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


