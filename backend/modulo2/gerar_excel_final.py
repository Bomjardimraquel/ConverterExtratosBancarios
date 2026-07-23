"""
Gerador do Excel final do Módulo 2 — reutilizável pra qualquer empresa.

- Aba "Lançamentos": só o que resultado_para_excel() devolve (nunca inclui
  'ja_lancado' nem pendências), com filtro automático no cabeçalho.
- Aba "Pendências": só quando existem — despesas/razão que não acharam
  correspondência no extrato (resultado_pendencias()).
- Aba "Legenda": explica as cores.
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

COR_HEADER = "5C5C5C"
COR_TEXTO_HEADER = "FFFFFF"
COR_CASADO = "F2F2F2"
COR_REVISAO = "FFEBEE"
COR_JA_BAIXADO = "FFF3CD"
COR_AJUSTAVEL = "E3F2FD"  # imposto/folha/aluguel — provavelmente certo, mas pode precisar de ajuste ou split manual

BORDA = Border(*[Side(style="thin", color="D9D9D9")] * 4)

ORIGEM_LEGIVEL = {
    "despesa": "Despesa", "titulo": "Título",
    "regra_texto": "Regra de texto", "simples": "Revisar",
}

# Traço usado como SEPARADOR DE FRASE (tem espaço dos dois lados, tipo
# "CDL - CDL MENSAL" ou "Casado — confiável") — vira dois-pontos, que é o
# que normalmente faz sentido nesse lugar (rótulo: explicação). Não mexe
# em traço colado dentro de número (tipo "0001-53" de CNPJ, ou código de
# conta), porque aí não tem espaço nenhum ao redor e o padrão não bate.
_TRAVESSAO_RE = re.compile(r"\s+[-–—]\s+")


def _remover_travessoes(texto) -> str:
    if not texto:
        return texto
    texto = _TRAVESSAO_RE.sub(": ", str(texto))
    return re.sub(r"\s{2,}", " ", texto).strip()


def _cabecalho(ws, titulo, headers, n_cols_titulo):
    col_fim = get_column_letter(n_cols_titulo)
    ws.merge_cells(f"A1:{col_fim}1")
    ws["A1"] = _remover_travessoes(titulo)
    ws["A1"].font = Font(bold=True, size=13, color=COR_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=COR_TEXTO_HEADER, size=10)
        cell.fill = PatternFill("solid", fgColor=COR_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDA
    ws.row_dimensions[2].height = 22


def gerar_excel_final(motor, resultado, titulo_planilha: str, caminho: str):
    para_excel = motor.resultado_para_excel(resultado)
    pendencias = motor.resultado_pendencias(resultado)

    wb = Workbook()

    # ── aba Lançamentos ─────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Lançamentos"
    headers = ["Data", "Descrição", "Tipo", "Débito", "Crédito",
               "Terceiro Débito", "Terceiro Crédito", "Valor (R$)", "Origem", "Observação"]
    _cabecalho(ws, titulo_planilha, headers, len(headers))

    row_idx = 3
    for r in para_excel:
        ja_baixado = "JÁ BAIXADO" in r.aviso
        if ja_baixado:
            fill = PatternFill("solid", fgColor=COR_JA_BAIXADO)
        elif r.confianca == "revisar_manual":
            fill = PatternFill("solid", fgColor=COR_AJUSTAVEL)
        elif r.origem == "simples":
            fill = PatternFill("solid", fgColor=COR_REVISAO)
        else:
            fill = PatternFill("solid", fgColor=COR_CASADO)

        vals = [
            r.data.strftime("%d/%m/%Y"), _remover_travessoes(r.descricao),
            "Crédito" if r.tipo == "C" else "Débito",
            r.conta_debito, r.conta_credito,
            r.terceiro_debito, r.terceiro_credito,
            r.valor, ORIGEM_LEGIVEL.get(r.origem, r.origem),
            _remover_travessoes(r.aviso) if not ja_baixado else "",
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = fill
            cell.border = BORDA
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col == 8:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col in (4, 5, 6, 7):
                cell.alignment = Alignment(horizontal="center")
        row_idx += 1

    total_row = row_idx
    ws.merge_cells(f"A{total_row}:G{total_row}")
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(bold=True, color=COR_TEXTO_HEADER)
    ws[f"A{total_row}"].fill = PatternFill("solid", fgColor=COR_HEADER)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="right")
    total_cell = ws.cell(row=total_row, column=8, value=f"=SUM(H3:H{total_row-1})")
    total_cell.font = Font(bold=True, color=COR_TEXTO_HEADER)
    total_cell.fill = PatternFill("solid", fgColor=COR_HEADER)
    total_cell.number_format = "#,##0.00"
    ws.cell(row=total_row, column=9).fill = PatternFill("solid", fgColor=COR_HEADER)
    ws.cell(row=total_row, column=10).fill = PatternFill("solid", fgColor=COR_HEADER)

    larguras = [12, 45, 10, 10, 10, 14, 14, 14, 16, 40]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg
    ws.freeze_panes = "A3"

    # filtro automático no cabeçalho (linha 2 até a última linha de dado,
    # sem incluir a linha de TOTAL)
    ws.auto_filter.ref = f"A2:J{row_idx - 1}"

    # ── aba Pendências (só se existir alguma) ───────────────────────────
    if pendencias:
        wsp = wb.create_sheet("Pendências")
        p_headers = ["Data", "Valor (R$)", "Tipo", "Descrição", "Aviso"]
        _cabecalho(wsp, "Itens sem correspondência (conferir manualmente)", p_headers, len(p_headers))
        for i, r in enumerate(pendencias, 3):
            fill = PatternFill("solid", fgColor=COR_REVISAO)
            vals = [r.data.strftime("%d/%m/%Y"), r.valor, "Crédito" if r.tipo == "C" else "Débito",
                    _remover_travessoes(r.descricao), _remover_travessoes(r.aviso)]
            for col, val in enumerate(vals, 1):
                cell = wsp.cell(row=i, column=col, value=val)
                cell.fill = fill
                if col == 2:
                    cell.number_format = "#,##0.00"
        wsp.column_dimensions['A'].width = 12
        wsp.column_dimensions['B'].width = 14
        wsp.column_dimensions['C'].width = 10
        wsp.column_dimensions['D'].width = 45
        wsp.column_dimensions['E'].width = 60
        wsp.auto_filter.ref = f"A2:E{2 + len(pendencias)}"

    # ── aba Legenda ──────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Legenda")
    ws2["A1"] = "Cor"; ws2["B1"] = "Significado"
    ws2["A1"].font = ws2["B1"].font = Font(bold=True)
    legendas = [
        (COR_CASADO, "Conciliado com despesa, título ou regra de texto (confiável, pode importar direto)"),
        (COR_AJUSTAVEL, "Imposto, folha ou aluguel: classificação provável, mas confira (pode precisar dividir entre contas, corrigir ou apagar)"),
        (COR_JA_BAIXADO, "Título já baixado no Prosoft (nome+valor+data batem): confirmar e descartar, não reimportar"),
        (COR_REVISAO, "Classificação simples (banco x caixa), sem terceiro: precisa revisar antes de importar"),
    ]
    for i, (cor, texto) in enumerate(legendas, 2):
        ws2.cell(row=i, column=1).fill = PatternFill("solid", fgColor=cor)
        ws2.cell(row=i, column=2, value=_remover_travessoes(texto))
    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 85

    wb.save(caminho)
    return {"linhas_lancamentos": row_idx - 3, "linhas_pendencias": len(pendencias)}