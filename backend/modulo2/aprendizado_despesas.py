"""
Aprendizado de classificação de despesas por Favorecido/Descrição.

Contexto: o escritório recebe um arquivo CRU de movimento contábil (com
colunas D. Previsão / D. Realização / D. Conciliação / V. Previsto /
V. Recebido / V. Realizado / Favorecido / Razão Social / Forma Pagamento /
Cartão Doc. / Cartão Aut. / Descrição / Observação — SEM Débito/Crédito),
e a Raquel classifica ele na mão todo mês, editando um arquivo com as
colunas Débito/Crédito preenchidas.

Esse módulo aprende com um mês já classificado (o "gabarito") e aplica a
mesma classificação automaticamente no mês seguinte, cru — usando
Favorecido (principal) e, quando o mesmo Favorecido aparece em contas
diferentes (ex: "WMS SUPERMERCADOS" ora é alimentação, ora é material de
limpeza), uma palavra-chave da Descrição pra desempatar.

O que NÃO dá pra aprender com confiança fica separado numa lista de
"não classificados" — pra revisão manual, igual já fazemos com título e
regra de texto sem match.
"""
import io
import re
import csv
import datetime
import unicodedata
from collections import defaultdict, Counter

from openpyxl import load_workbook

try:
    from .modelos import Despesa
    from .cruzamento import norm, contem_termo
except ImportError:
    from modelos import Despesa
    from cruzamento import norm, contem_termo


def _norm(s) -> str:
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    return " ".join(s.upper().split())


def _norm_cabecalho(s: str) -> str:
    """Normaliza nome de coluna pra comparar sem depender de acento,
    ponto ou espaço extra ('D. Realização' vira 'D REALIZACAO')."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode()
    s = re.sub(r"[.\-_/]", " ", s)
    return " ".join(s.upper().split())


def _converter_celula_csv(valor: str):
    """CSV/TSV só tem texto — tenta converter pro tipo real (data ou
    número), igual o openpyxl já devolve nativamente pra .xlsx. Deixa
    como string se não parecer nem data nem número."""
    if valor is None:
        return None
    texto = valor.strip()
    if texto == "":
        return None
    m = re.match(r"^(\d{2})/(\d{2})/(\d{4})$", texto)
    if m:
        d, mes, a = m.groups()
        try:
            return datetime.date(int(a), int(mes), int(d))
        except ValueError:
            return texto
    # número no formato brasileiro (vírgula decimal, sem separador de milhar
    # ambíguo aqui já que os valores de despesa não costumam ter milhar)
    if re.match(r"^-?\d+(,\d+)?$", texto):
        try:
            return float(texto.replace(",", "."))
        except ValueError:
            return texto
    return texto


def _abrir_planilha(caminho_ou_bytes):
    """
    Devolve (cabecalho, linhas) — lista de tuplas, já sem o cabeçalho.
    Aceita tanto .xlsx de verdade (ZIP) quanto CSV/TSV de texto puro (o
    caso real: a Arandu manda às vezes um arquivo com extensão .csv que,
    por dentro, é texto separado por tabulação — abre no Excel por
    associação do Windows, mas não é um .xlsx de verdade).
    """
    if isinstance(caminho_ou_bytes, (bytes, bytearray)):
        conteudo = caminho_ou_bytes
    else:
        with open(caminho_ou_bytes, "rb") as f:
            conteudo = f.read()

    if conteudo[:4] == b"PK\x03\x04":
        wb = load_workbook(io.BytesIO(conteudo), data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        return linhas[0], linhas[1:]

    # Não é ZIP — trata como texto (CSV ou TSV). Detecta o separador
    # olhando a primeira linha: tabulação é o mais comum nesses exports,
    # mas aceita vírgula ou ponto e vírgula também.
    for encoding in ("utf-8-sig", "utf-8", "windows-1252", "latin-1"):
        try:
            texto = conteudo.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        raise ValueError("Não consegui decodificar o arquivo CSV (codificação desconhecida).")

    primeira_linha = texto.splitlines()[0] if texto.splitlines() else ""
    contagens = {"\t": primeira_linha.count("\t"), ";": primeira_linha.count(";"), ",": primeira_linha.count(",")}
    separador = max(contagens, key=contagens.get)

    leitor = csv.reader(io.StringIO(texto), delimiter=separador)
    todas_linhas = [tuple(row) for row in leitor if any(c.strip() for c in row)]
    if not todas_linhas:
        raise ValueError("Arquivo CSV está vazio.")

    cabecalho = todas_linhas[0]
    linhas = [tuple(_converter_celula_csv(c) for c in row) for row in todas_linhas[1:]]
    return cabecalho, linhas


def _indice_colunas(cabecalho, nomes_esperados: dict) -> dict:
    """
    Acha o índice de cada coluna pelo NOME (não pela posição fixa) — o
    arquivo real pode ter coluna a mais/reordenada (ex: a Arandu manda um
    CSV com "D. Emissão" no meio, que o formato original de 13 colunas
    não tinha). `nomes_esperados` é {chave_interna: [aliases possíveis]}.
    Levanta ValueError se alguma coluna OBRIGATÓRIA não for encontrada.
    """
    cabecalho_norm = [_norm_cabecalho(c) for c in cabecalho]
    indices = {}
    for chave, aliases in nomes_esperados.items():
        idx = None
        for alias in aliases:
            alias_norm = _norm_cabecalho(alias)
            if alias_norm in cabecalho_norm:
                idx = cabecalho_norm.index(alias_norm)
                break
        indices[chave] = idx
    return indices


def treinar_classificador_despesas(caminho_ou_bytes_classificado) -> dict:
    """
    Lê o arquivo JÁ CLASSIFICADO (com Débito/Crédito preenchidos) e monta
    o "modelo" de aprendizado. Colunas esperadas (na ordem, a partir da
    coluna A): D. Realização | Razão Social | Descrição | V. Realizado |
    Débito | Crédito — mesma estrutura de sempre pro arquivo classificado.

    Devolve um dict:
      {
        "favorecido_unico": {favorecido_normalizado: conta_debito, ...},
        "favorecido_ambiguo": {
            favorecido_normalizado: [(palavra_chave, conta_debito), ...]
        },
      }
    """
    _cabecalho, linhas = _abrir_planilha(caminho_ou_bytes_classificado)

    contas_por_favorecido = defaultdict(list)  # favorecido -> [(descricao, conta_debito), ...]
    for row in linhas:
        if not row or row[0] is None:
            continue
        razao, descricao, valor, deb = row[1], row[2], row[3], row[4]
        favorecido = _norm(razao)
        if not favorecido or deb is None:
            continue
        contas_por_favorecido[favorecido].append((_norm(descricao), deb))

    favorecido_unico = {}
    favorecido_ambiguo = {}

    for favorecido, ocorrencias in contas_por_favorecido.items():
        contas_distintas = {c for _, c in ocorrencias}
        if len(contas_distintas) == 1:
            favorecido_unico[favorecido] = ocorrencias[0][1]
            continue

        # Favorecido ambíguo (mais de uma conta) — acha palavra(s) na
        # descrição que aparecem SÓ nas ocorrências de uma conta e não
        # nas outras, pra usar como critério de desempate.
        palavras_por_conta = defaultdict(Counter)
        for descricao, conta in ocorrencias:
            for palavra in set(descricao.split()):
                if len(palavra) > 2:
                    palavras_por_conta[conta][palavra] += 1

        regras = []
        todas_contas = list(palavras_por_conta.keys())
        for conta in todas_contas:
            outras = set()
            for outra_conta in todas_contas:
                if outra_conta != conta:
                    outras |= set(palavras_por_conta[outra_conta])
            exclusivas = set(palavras_por_conta[conta]) - outras
            if exclusivas:
                # usa a palavra exclusiva mais frequente pra essa conta
                palavra_escolhida = palavras_por_conta[conta].most_common(1)[0][0]
                regras.append((palavra_escolhida, conta))

        if regras:
            # Quando só existem 2 contas possíveis pra esse favorecido e
            # achamos palavra-chave pra UMA delas, a outra vira o "padrão"
            # (usada quando a palavra-chave não aparece) — em vez de
            # deixar sem classificar sempre que a palavra não bater.
            # Ex: WMS Supermercados — "LIMPEZA" na descrição -> material
            # de limpeza; sem essa palavra -> alimentação (o caso comum).
            if len(todas_contas) == 2 and len(regras) == 1:
                conta_com_palavra = regras[0][1]
                conta_padrao = next(c for c in todas_contas if c != conta_com_palavra)
                favorecido_ambiguo[favorecido] = {
                    "regras": regras,
                    "padrao": conta_padrao,
                }
            else:
                favorecido_ambiguo[favorecido] = {"regras": regras, "padrao": None}
        # se não achou nenhuma palavra que distingue, esse favorecido some
        # do modelo — melhor não classificar do que classificar errado
        # (fica pra revisão manual quando aparecer no arquivo cru)

    return {
        "favorecido_unico": favorecido_unico,
        "favorecido_ambiguo": favorecido_ambiguo,
    }


def _classificar_uma_despesa(favorecido_norm: str, descricao_norm: str, modelo: dict):
    """Devolve a conta_debito aprendida, ou None se não souber classificar."""
    if favorecido_norm in modelo["favorecido_ambiguo"]:
        info = modelo["favorecido_ambiguo"][favorecido_norm]
        for palavra_chave, conta in info["regras"]:
            if palavra_chave in descricao_norm.split():
                return conta
        # nenhuma palavra-chave bateu — usa o padrão (conta mais comum
        # pra esse favorecido), se existir; senão manda pra revisão manual
        return info["padrao"]

    return modelo["favorecido_unico"].get(favorecido_norm)


def _tentar_regras_texto(descricao: str, razao_social: str, regras_texto: list):
    """
    Segunda tentativa de classificação: as MESMAS regras_texto que já
    existem na config da empresa pra casar lançamento do banco (SALARIO,
    REPASSE, ALUGUEL, IRRF/FGTS/PIS/COFINS etc) — reaproveitadas aqui pra
    classificar a despesa quando o aprendizado por Favorecido não sabe
    (comum pra pagamento de pessoa física com nome variado: cada
    funcionária tem um nome diferente, mas todas batem em "SALARIO").
    Devolve a conta_debito, ou None se nenhuma regra bateu.
    """
    if not regras_texto:
        return None
    texto = norm(f"{razao_social} {descricao}")
    for regra in regras_texto:
        if regra.get("usa_padrao"):
            continue  # regra de descrição, não se aplica aqui (não tem conta fixa)
        if regra.get("contem_todos"):
            bate = all(contem_termo(texto, p) for p in regra["contem_todos"])
        else:
            bate = any(contem_termo(texto, p) for p in regra.get("contem", []))
        if not bate:
            continue
        if any(contem_termo(texto, p) for p in regra.get("nao_contem", [])):
            continue
        conta = regra.get("debito_D")
        if conta:
            return conta
    return None


_MESES_ABREV = {
    1: "JAN", 2: "FEV", 3: "MAR", 4: "ABR", 5: "MAI", 6: "JUN",
    7: "JUL", 8: "AGO", 9: "SET", 10: "OUT", 11: "NOV", 12: "DEZ",
}


def _extrair_mes_ano(descricao: str) -> str:
    """Extrai 'MM AAAA' ou 'MM.AAAA' da descrição (ex: 'PRO LABORE 04 2026')
    e devolve formatado tipo 'ABR/2026'. Devolve '' se não achar."""
    m = re.search(r"\b(\d{1,2})[.\s](\d{4})\b", descricao or "")
    if not m:
        return ""
    mes, ano = int(m.group(1)), m.group(2)
    if 1 <= mes <= 12:
        return f"{_MESES_ABREV[mes]}/{ano}"
    return ""


_SUFIXOS_PJ_RE = re.compile(
    r"\bLTDA\b|\bLTD\b|\bEIRELI\b|\bS[/\.]?A\b|\bCIA\b|\bCOOP\b|\bASSOC|\bME\b|\bEPP\b|"
    r"\bCOM(ERCIO)?\b|\bIND(USTRIA)?\b|\bDISTRIB|\bIMPORT|\bEXPORT|\bSERVICOS?\b|"
    r"\bASSESSORIA\b|\bCONSULTORIA\b|\bTRANSPORTES?\b|\bLOGISTICA\b|\bFEDERAL\b|"
    r"\bMUNICIPIO\b|\bPREFEITURA\b|\bSECRETARIA\b|\bCAIXA\b|\bBANCO\b|\bCOOPERATIVA\b",
    re.IGNORECASE
)


def _parece_pessoa_fisica(nome: str) -> bool:
    """Heurística: nome de 2-5 palavras, sem sufixo de empresa/órgão
    público, sem número — mesma lógica usada antes pra separar PF/PJ nos
    balancetes."""
    nome = (nome or "").strip()
    if not nome or _SUFIXOS_PJ_RE.search(nome):
        return False
    palavras = nome.split()
    return 2 <= len(palavras) <= 5 and not any(ch.isdigit() for ch in nome)


def _tentar_regras_extras(favorecido: str, razao_social: str, descricao: str, regras_extras: list):
    """
    Regras específicas DESSE arquivo de movimento contábil (não as
    regras_texto do extrato) — categorias como alimentação/material de
    consumo/prestação de serviço PF, e casos com formatação de descrição
    própria (ex: pró-labore, CDL, ISSQN). Cada regra pode ter:
      "contem": [...]            -> palavras que disparam a regra
      "conta_debito": "..."
      "confianca": "..." (opcional, default "media")
      "aviso": "..." (opcional)
      "descricao_template": "..." (opcional) -> pode usar {favorecido},
          {razao_social}, {descricao}, {mes_ano}
    Devolve (conta_debito, descricao_formatada, confianca, aviso) ou
    (None, None, None, None) se nenhuma regra bateu.
    """
    if not regras_extras:
        return None, None, None, None
    texto = norm(f"{favorecido} {razao_social} {descricao}")
    for regra in regras_extras:
        if not any(contem_termo(texto, p) for p in regra.get("contem", [])):
            continue
        conta = regra.get("conta_debito")
        template = regra.get("descricao_template")
        if template:
            mes_ano = _extrair_mes_ano(descricao)
            desc_formatada = template.format(
                favorecido=favorecido, razao_social=razao_social,
                descricao=descricao, mes_ano=mes_ano,
            )
        else:
            desc_formatada = None
        return conta, desc_formatada, regra.get("confianca", "media"), regra.get("aviso", "")
    return None, None, None, None


def carregar_despesas_brutas(
    caminho_ou_bytes, modelo: dict, regras_texto: list = None,
    regras_extras: list = None, ignorar_se_contem: list = None,
) -> tuple:
    """
    Lê o arquivo CRU (sem Débito/Crédito) e aplica o modelo aprendido.

    Colunas esperadas (na ordem): D. Previsão | D. Realização |
    D. Conciliação | V. Previsto | V. Recebido | V. Realizado |
    Favorecido | Razão Social | Forma Pagamento | Cartão Doc. |
    Cartão Aut. | Descrição | Observação

    Regra de negócio: linha com "V. Recebido" preenchido é recebimento
    (receita), não despesa — é ignorada. Usa "V. Realizado" como valor.

    Devolve (despesas_classificadas, nao_classificadas):
      - despesas_classificadas: list[Despesa], prontas pro motor de
        cruzamento (conta_credito fica None de propósito — o motor já
        sabe usar a conta do banco como padrão quando não vem preenchida)
      - nao_classificadas: list de dicts (favorecido, descricao, valor,
        data, forma_pagamento) — Favorecido/Descrição sem match no
        modelo aprendido, precisa de olho humano
    """
    NOMES_ESPERADOS = {
        "d_realizacao": ["D. Realização", "D Realizacao"],
        "v_recebido": ["V. Recebido", "V Recebido"],
        "v_realizado": ["V. Realizado", "V Realizado"],
        "favorecido": ["Favorecido"],
        "razao_social": ["Razão Social", "Razao Social"],
        "forma_pagamento": ["Forma Pagamento"],
        "descricao": ["Descrição", "Descricao"],
    }
    cabecalho, linhas = _abrir_planilha(caminho_ou_bytes)
    idx = _indice_colunas(cabecalho, NOMES_ESPERADOS)
    faltando = [k for k in ("d_realizacao", "v_realizado", "favorecido") if idx[k] is None]
    if faltando:
        raise ValueError(
            f"Não achei a(s) coluna(s) {faltando} no cabeçalho do arquivo de movimento bruto. "
            f"Cabeçalho encontrado: {list(cabecalho)}"
        )

    def val(row, chave):
        i = idx[chave]
        return row[i] if i is not None and i < len(row) else None

    despesas_classificadas = []
    nao_classificadas = []

    for row in linhas:
        d_realizacao = val(row, "d_realizacao")
        if d_realizacao is None:  # sem D. Realização, ignora
            continue
        v_recebido = val(row, "v_recebido")
        v_realizado = val(row, "v_realizado")
        favorecido = val(row, "favorecido")
        razao_social = val(row, "razao_social")
        forma_pagamento = val(row, "forma_pagamento")
        descricao = val(row, "descricao")

        if v_recebido is not None:
            continue  # é recebimento, não despesa — ignora
        if v_realizado is None:
            continue  # nada realizado ainda, não interessa pro cruzamento

        texto_ignorar = norm(f"{favorecido} {razao_social} {descricao}")
        if ignorar_se_contem and any(contem_termo(texto_ignorar, p) for p in ignorar_se_contem):
            continue  # ex: SALARIO — já é tratado via regra_texto do lado do extrato

        data = d_realizacao.date() if hasattr(d_realizacao, "date") else d_realizacao
        favorecido_norm = _norm(favorecido)
        # Razão Social costuma ser o nome mais completo/estável — prioriza
        # ela pra bater com o modelo (que foi treinado com Razão Social);
        # cai pro Favorecido (mais curto, tipo "PRO-RAD") se não tiver.
        chave_norm = _norm(razao_social) or favorecido_norm
        descricao_norm = _norm(descricao)

        conta_debito = _classificar_uma_despesa(chave_norm, descricao_norm, modelo)
        descricao_override = None
        confianca = None
        aviso = ""

        if conta_debito is None:
            conta_debito = _tentar_regras_texto(descricao or "", razao_social or "", regras_texto)

        if conta_debito is None:
            conta_debito, descricao_override, confianca, aviso = _tentar_regras_extras(
                favorecido or "", razao_social or "", descricao or "", regras_extras
            )

        if conta_debito is None and _parece_pessoa_fisica(razao_social or favorecido or ""):
            conta_debito = "53956"
            confianca = "revisar_manual"
            aviso = "Classificado como serviço de pessoa física só pelo nome (parece pessoa) — confirmar"

        if conta_debito is not None:
            despesas_classificadas.append(Despesa(
                data=data,
                razao_social=razao_social or favorecido or "",
                descricao=descricao or "",
                valor=float(v_realizado),
                conta_debito=conta_debito,
                conta_credito=None,  # o motor usa a conta do banco como padrão
                confianca=confianca,
                aviso=aviso,
                descricao_override=descricao_override,
            ))
        else:
            nao_classificadas.append({
                "data": str(data),
                "favorecido": favorecido or "",
                "razao_social": razao_social or "",
                "descricao": descricao or "",
                "valor": float(v_realizado),
                "forma_pagamento": forma_pagamento or "",
            })

    return despesas_classificadas, nao_classificadas