"""
Detecção automática de tipo de arquivo de despesa — substitui o campo
`modo_despesas` fixo na config da empresa.

Antes: a config dizia "casar" ou "desconsiderar" pra cada empresa (nomes
que descreviam a AÇÃO do motor, não o tipo do arquivo), e o motor confiava
nisso. Problema: se um mês a empresa mandar um arquivo diferente do que
sempre mandou (ou errar o upload), o motor processa do jeito errado sem
avisar.

Agora: decide pela ESTRUTURA do próprio arquivo, sempre, pra qualquer
empresa. Três formatos possíveis:
  - "despesa_classificada": .xlsx de verdade, 6 colunas (Data / Razão
    Social / Descrição / Valor / Débito / Crédito) — já vem com a conta
    escolhida na mão.
  - "razao_prosoft": SpreadsheetML — um XML disfarçado de .xls (o Prosoft
    usa a extensão .xls, mas o conteúdo é texto XML, não o formato
    binário/zip real de Excel).
  - "movimento_bruto": .xlsx de verdade também, mas com 13 colunas
    (D. Previsão / D. Realização / ... / Favorecido / Razão Social /
    Forma Pagamento / ... / Descrição / Observação) — SEM Débito/Crédito.
    Esse é o arquivo cru que o escritório recebe todo mês; precisa do
    aprendizado_despesas.py (+ um mês já classificado de referência) pra
    virar despesa classificada.

Como "despesa_classificada" e "movimento_bruto" são os dois arquivos ZIP
de verdade (mesma assinatura de bytes), a distinção entre os dois só dá
pra fazer olhando o CABEÇALHO por dentro — "Favorecido" só existe no
formato bruto.
"""
import io
from openpyxl import load_workbook

try:
    from .modelos import Despesa
    from .parser_razao import parse_razao_ja_lancado
except ImportError:
    from modelos import Despesa
    from parser_razao import parse_razao_ja_lancado


def _eh_movimento_bruto(conteudo: bytes) -> bool:
    """Espia o cabeçalho do xlsx: se tiver uma coluna 'Favorecido', é o
    formato cru (movimento_bruto), não o classificado."""
    try:
        wb = load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
        ws = wb.active
        primeira_linha = next(ws.iter_rows(max_row=1, values_only=True), ())
        cabecalho = " ".join(str(c or "") for c in primeira_linha).upper()
        return "FAVORECIDO" in cabecalho
    except Exception:
        return False


def detectar_tipo_arquivo(conteudo: bytes) -> str:
    """
    Devolve 'despesa_classificada', 'movimento_bruto' ou 'razao_prosoft'
    — nomeados pelo TIPO do arquivo, não pela ação que o motor faz com
    ele (o antigo `modo_despesas` usava "casar"/"desconsiderar", que
    descreviam a ação e exigiam conhecer o código pra entender o que
    representavam).

    Levanta ValueError se não conseguir identificar nenhum dos formatos
    — nesse caso o arquivo deve ser rejeitado no upload, não adivinhado.
    """
    # .xlsx de verdade é um arquivo ZIP — começa sempre com esses 4 bytes
    if conteudo[:4] == b"PK\x03\x04":
        return "movimento_bruto" if _eh_movimento_bruto(conteudo) else "despesa_classificada"

    # Razão do Prosoft: XML disfarçado de .xls. Tenta decodificar como
    # texto e procura a assinatura do SpreadsheetML.
    for encoding in ("windows-1252", "utf-8"):
        try:
            inicio = conteudo[:2000].decode(encoding, errors="ignore")
        except Exception:
            continue
        if "schemas-microsoft-com:office:spreadsheet" in inicio or "<?xml" in inicio[:100]:
            return "razao_prosoft"

    raise ValueError(
        "Não consegui identificar o tipo do arquivo (não parece xlsx nem "
        "razão SpreadsheetML do Prosoft). Confirma se o arquivo certo foi enviado."
    )


def _parse_data_texto(valor) -> "date | None":
    """Célula de data que veio como texto (ex: '01/05/2026') em vez de data
    de verdade — acontece quando a planilha não está formatada como data.
    Devolve None se não conseguir interpretar (a linha é descartada, não
    quebra o carregamento inteiro)."""
    import datetime
    if valor is None:
        return None
    txt = str(valor).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.datetime.strptime(txt, fmt).date()
        except ValueError:
            continue
    return None


def _carregar_despesas_xlsx(caminho_ou_bytes) -> list:
    """
    Lê o arquivo de despesa classificada. Aceita caminho de arquivo ou
    bytes. Valida a ESTRUTURA (tipo de dado de cada coluna) antes de
    processar, em vez do texto do cabeçalho — o nome das colunas varia
    entre formatos reais (o que a Arandu manda de verdade usa "D.
    Realização"/"V. Realizado", não "Data"/"Valor" — mesma coisa, nome
    diferente), então validar por texto exato quebra em qualquer variação
    de nomenclatura. A ordem das 6 colunas (data, razão social, descrição,
    valor, débito, crédito) é o que realmente importa e não muda.
    """
    origem = io.BytesIO(caminho_ou_bytes) if isinstance(caminho_ou_bytes, (bytes, bytearray)) else caminho_ou_bytes
    wb = load_workbook(origem, data_only=True)
    ws = wb.active  # primeira aba, não assume nome fixo tipo "Planilha1"

    linhas = ws.iter_rows(values_only=True)
    cabecalho = next(linhas, None)
    if cabecalho is None:
        raise ValueError("Arquivo de despesas está vazio.")

    # Acha a primeira linha de dado de verdade (pula linhas vazias entre o
    # cabeçalho e o início dos dados, se houver) pra validar a estrutura.
    linhas_restantes = []
    primeira_linha_dado = None
    for row in linhas:
        linhas_restantes.append(row)
        if primeira_linha_dado is None and row and len(row) >= 6 and row[0] is not None:
            primeira_linha_dado = row

    if primeira_linha_dado is None:
        raise ValueError("Arquivo de despesas não tem nenhuma linha de dado (só cabeçalho?).")

    col_data, _col_razao, _col_desc, col_valor, col_deb, col_cred = primeira_linha_dado[:6]
    data_ok = hasattr(col_data, "date") or _parse_data_texto(col_data) is not None
    valor_ok = isinstance(col_valor, (int, float))
    contas_ok = all(
        isinstance(c, (int, float)) or (isinstance(c, str) and c.strip().replace(".", "").isdigit())
        for c in (col_deb, col_cred) if c is not None
    )
    if not (data_ok and valor_ok and contas_ok):
        raise ValueError(
            "Estrutura do arquivo de despesas não bate com o esperado "
            "(esperava: data, razão social, descrição, valor numérico, "
            f"conta débito, conta crédito). Primeira linha de dado: {primeira_linha_dado[:6]}"
        )

    linhas = linhas_restantes

    despesas = []
    for row in linhas:
        if len(row) < 6:
            continue
        data, razao, desc, valor, deb, cred = row[:6]
        if data is None or valor is None:
            continue
        data_convertida = data.date() if hasattr(data, "date") else _parse_data_texto(data)
        if data_convertida is None:
            continue
        despesas.append(Despesa(
            data=data_convertida,
            razao_social=razao or "",
            descricao=desc or "",
            valor=float(valor),
            conta_debito=deb,
            conta_credito=cred,
        ))
    return despesas


def carregar_arquivo_despesas_ou_razao(caminho_ou_bytes):
    """
    Ponto de entrada único. Devolve uma tupla (modo_despesas, dados):
      - ("despesa_classificada", list[Despesa])     — despesa classificada
      - ("razao_prosoft", list[DespesaJaLancada])   — razão do Prosoft
      - ("movimento_bruto", bytes)                  — arquivo cru, sem
        classificação; devolve os BYTES originais porque esse formato
        precisa do aprendizado_despesas.py + um arquivo de referência já
        classificado, que essa função não tem acesso — quem chama decide
        o que fazer (ver tasks_modulo2.py).

    Uso típico na rota/integração:

        modo, dados = carregar_arquivo_despesas_ou_razao(caminho)
        if modo == "despesa_classificada":
            resultado = motor.cruzar(lancamentos, despesas=dados, modo_despesas="despesa_classificada")
        elif modo == "razao_prosoft":
            resultado = motor.cruzar(lancamentos, despesas_ja_lancadas=dados, modo_despesas="razao_prosoft")
        else:  # "movimento_bruto" — ver tasks_modulo2.py pro fluxo completo
            ...
    """
    if isinstance(caminho_ou_bytes, (bytes, bytearray)):
        conteudo = caminho_ou_bytes
    else:
        with open(caminho_ou_bytes, "rb") as f:
            conteudo = f.read()

    modo = detectar_tipo_arquivo(conteudo)

    if modo == "razao_prosoft":
        # parse_razao_ja_lancado espera um CAMINHO (abre o arquivo ela
        # mesma) — se recebemos bytes, salva num temporário primeiro.
        if isinstance(caminho_ou_bytes, (bytes, bytearray)):
            import tempfile
            with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
                tmp.write(conteudo)
                caminho_tmp = tmp.name
            return "razao_prosoft", parse_razao_ja_lancado(caminho_tmp)
        return "razao_prosoft", parse_razao_ja_lancado(caminho_ou_bytes)

    if modo == "movimento_bruto":
        return "movimento_bruto", conteudo

    # Sempre passa os BYTES (não o caminho) daqui pra frente — o openpyxl
    # rejeita qualquer caminho terminando em ".xls" só pela extensão do
    # nome do arquivo (nem olha o conteúdo), mesmo quando por dentro é um
    # .xlsx de verdade (caso real: a Arandu manda a despesa classificada
    # com extensão .xls, mas o conteúdo é xlsx/zip). Passando bytes em vez
    # do caminho, esse cheque de extensão nunca entra em ação.
    return "despesa_classificada", _carregar_despesas_xlsx(conteudo)